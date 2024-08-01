# import a package to open up an excel file
import openpyxl

workbook = openpyxl.load_workbook("transactions.xlsx")
sheet = workbook["Sheet1"]
# If you want to grab information from a singular cell, you can grab a cell 1 of 2 ways
# 1. Use square bracket notation with the Col and Row as a string
cell = sheet["a1"]
# 2. Use the cell(row, col) method on the sheet object passing
cell2 = sheet.cell(1, 2)

print(cell.value)  # returns value at cell A1: transaction_id
print(cell2.value)  # returns value at cell B1: product_id

# Let's get 90% of the value listed in every row's col 3
# steps:
""" 
1. determine how many rows the sheet has
2. make a loop to iterate through all the rows
3. get every row's value from the 3rd column
4. calculate 90% of the value in the cells to get the corrected price
5. Add a new column to the sheet to store the corrected prices
6. Set the corrected price value into the corrected price column
7. Save the changes made to the sheet in a new file incase there's a bug in the code
"""
# Lets a a chart/graph/visualization to the sheet
"""
A. import barchart and reference from openyxl.chart  (imports should be at top of the page)
B. grab the corrected values from a column 4
C. create a chart using the corrected values
D. Add the new chart to the sheet. method is sheet.add_chart(chart, 'location' )
E. 
"""
from openpyxl.chart import BarChart, Reference

row_count = sheet.max_row  # Step 1: there are four rows
# print(row_count)

for row in range(
    2, row_count + 1
):  # Step 2: we start from 2 to exclude the title & +1 to include the last number,
    cell = sheet.cell(row, 3)  # Step 3
    corrected_price = cell.value * 0.9  # Step 4
    corrected_price_cell = sheet.cell(row, 4)  # Step 5
    corrected_price_cell.value = corrected_price  # Step 6
    print(corrected_price)

values = Reference(
    sheet,
    min_row=2,
    max_row=row_count,  # Step B - args: (sheet used, StartRow, EndRow, StartCol, EndCol)
    min_col=4,
    max_col=4,
)

chart = BarChart()
chart.add_data(values)  # Step C
sheet.add_chart(chart, 'e2') # Step D

workbook.save("transactions2.xlsx")  # Step 7
