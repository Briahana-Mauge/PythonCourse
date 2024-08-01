# Correct the prices of this spreadsheet to reflect a 30% discount and make a chart of the new prices

import openpyxl
from openpyxl.chart import BarChart, Reference
from pathlib import Path


def process_workbook(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Sheet1"]
    cell = sheet["a1"]

    row_count = sheet.max_row


    for row in range(
        2, row_count + 1
    ): 
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.7 
        corrected_price_cell = sheet.cell(row, 4) 
        corrected_price_cell.value = corrected_price  
        print(corrected_price)

    values = Reference(
        sheet,
        min_row=2,
        max_row=row_count,
        min_col=4,
        max_col=4,
    )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    workbook.save(filename) 


path = Path()


for file in path.glob('transactions2.xlsx'):
    process_workbook(file)
        
